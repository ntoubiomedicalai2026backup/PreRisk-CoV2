#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
PreRisk-CoV2: Pre-exposure Risk Assessment for SARS-CoV-2
==========================================================
Development and External Validation of a Pre-Exposure Protein Biomarker Panel and Machine Learning Model for Predicting SARS-CoV-2 Infection Risk

Author: NTOU Biomedical AI LAB
GitHub: https://github.com/NTOUBiomedicalAILAB/PreRisk-CoV2
"""

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import argparse
import os
import sys
import time
import pickle
from math import sqrt
from sklearn import preprocessing
from sklearn.metrics import (confusion_matrix, roc_curve, roc_auc_score,
                             precision_recall_curve, average_precision_score)
from sklearn.model_selection import LeaveOneOut
from sklearn.neighbors import KNeighborsClassifier
from imblearn.over_sampling import SMOTE
import openpyxl
import tenseal as ts
import warnings
warnings.filterwarnings('ignore')

###############################################################################
# BANNER & PROGRESS BAR
###############################################################################

def print_banner():
    banner = """
    ╔═══════════════════════════════════════════════════════════════╗
    ║                      PreRisk-CoV2                             ║
    ║         SARS-CoV-2 Pre-exposure Risk Assessment               ║
    ║          KNN-GA Protein Biomarker Framework                   ║
    ╚═══════════════════════════════════════════════════════════════╝
    """
    print(banner)


def print_progress_bar(iteration, total, prefix='Progress', suffix='Complete', length=40, fill='█'):
    """Clean progress bar to hide lengthy training and encryption processes"""
    percent = ("{0:.1f}").format(100 * (iteration / float(total)))
    filled_length = int(length * iteration // total)
    bar = fill * filled_length + '-' * (length - filled_length)
    sys.stdout.write(f'\r  {prefix} |{bar}| {percent}% {suffix}')
    sys.stdout.flush()
    if iteration == total:
        print()


###############################################################################
# SHARED UTILITIES (FHE Module Utilities)
###############################################################################

DEFAULT_PRIVATE_CTX  = 'private_context.bin'
DEFAULT_PUBLIC_CTX   = 'public_context.bin'
DEFAULT_ENC_TRAIN    = 'Discovery_encrypted.bin'
DEFAULT_TRAIN_LABELS = 'train_labels.pkl'

def load_context(path: str, require_private: bool = False) -> ts.Context:
    if not os.path.isfile(path):
        raise FileNotFoundError(f'[ERROR] Key file not found: {path!r}')
    with open(path, 'rb') as f:
        ctx = ts.context_from(f.read())
    if require_private and not ctx.is_private():
        raise ValueError(f'[CRITICAL ERROR] Private Context is required for this operation!')
    return ctx

def load_bin(path: str) -> dict:
    if not os.path.isfile(path):
        raise FileNotFoundError(f'[ERROR] Data file not found: {path!r}')
    with open(path, 'rb') as f:
        return pickle.load(f)

def save_bin(path: str, payload: dict):
    os.makedirs(os.path.dirname(os.path.abspath(path)) or '.', exist_ok=True)
    with open(path, 'wb') as f:
        pickle.dump(payload, f, protocol=pickle.HIGHEST_PROTOCOL)
    print(f'\n[INFO] Encrypted result file saved → {path}')


###############################################################################
# DATA PROCESSING 
###############################################################################

def missing_counts(data):
    missing = data.isnull().sum()
    missing = missing[missing > 0]
    missing.sort_values(inplace=True)
    df_mc = pd.DataFrame({'ColumnName': missing.index,
                          'MissingCount': missing.values})
    df_mc['Percentage(%)'] = df_mc['MissingCount'].apply(
        lambda x: round(x / data.shape[0] * 100, 2))
    return df_mc

def data_processing(df, scaler=None):
    sample_id  = df['sample ID'].values
    protein_id = df.columns.tolist()[2:94]   

    df_proc = df.drop(['sample ID'], axis=1)
    df_proc['PCR result'] = df_proc['PCR result'].map(
        {'Not': 0, 'Detected': 1}).astype(int)

    ndarray  = df_proc.values
    label    = ndarray[:, 0]
    features = ndarray[:, 1:]

    if scaler is None:
        scaler   = preprocessing.MinMaxScaler(feature_range=(0, 1))
        features = scaler.fit_transform(features)
    else:
        features = scaler.transform(features)

    return sample_id, protein_id, features, label, scaler


###############################################################################
# KNN BUILDER
###############################################################################

def build_knn_model(n_neighbors=5, leaf_size=30, algorithm='auto',
                    weights='uniform', p=2):
    return KNeighborsClassifier(
        n_neighbors=n_neighbors,
        leaf_size=leaf_size,
        algorithm=algorithm,
        weights=weights,
        p=p,
    )


###############################################################################
# ROC / PR CURVES
###############################################################################

def plot_roc_pr_curves(y_true, y_proba, save_path=None):
    fig = plt.gcf()
    fig.set_size_inches(16, 6)

    ax1 = plt.subplot(121)
    ax1.set_box_aspect(1)
    dnn_auc            = roc_auc_score(y_true, y_proba)
    dnn_fpr, dnn_tpr, _ = roc_curve(y_true, y_proba)
    plt.plot(dnn_fpr, dnn_tpr, marker='.',
             label=' (AUROC = %0.3f)' % dnn_auc)
    plt.plot([0, 1], [0, 1], color='red', linestyle='--')
    plt.fill_between(dnn_fpr, dnn_tpr, color='gray', alpha=0.2)
    plt.xlim([0.0, 1.0]); plt.ylim([0.0, 1.0])
    plt.title('ROC')
    plt.xlabel('False Positive Rate')
    plt.ylabel('True Positive Rate')
    plt.legend(loc='lower right')

    ax2 = plt.subplot(122)
    ax2.set_box_aspect(1)
    precision_v, recall_v, _ = precision_recall_curve(y_true, y_proba)
    dnn_aup = average_precision_score(y_true, y_proba)
    plt.plot(recall_v, precision_v, marker='.',
             label=' (AUPRC = %0.3f)' % dnn_aup)
    plt.fill_between(recall_v, precision_v, color='gray', alpha=0.2)
    plt.xlim([0.0, 1.0]); plt.ylim([0.0, 1.0])
    plt.title('PR')
    plt.xlabel('Recall', fontsize=14)
    plt.ylabel('Precision', fontsize=14)
    plt.legend(loc='lower right')

    plt.tight_layout()
    if save_path:
        os.makedirs(os.path.dirname(os.path.abspath(save_path)), exist_ok=True)
        plt.savefig(save_path, dpi=300, bbox_inches='tight')
    plt.clf()


###############################################################################
# EXCEL SAVE — INTERNAL / EXTERNAL
###############################################################################

def save_internal_results(data_save, loop, output_path, sheet_name):
    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    try:
        wb = openpyxl.load_workbook(output_path)
    except FileNotFoundError:
        wb = openpyxl.Workbook()

    wb.create_sheet(sheet_name, 0)
    ws = wb.active

    ws.append(['Results of each Cross-Validation iteration:'])
    ws.append([''])
    ws.append([' ', 'Accuracy', 'Specificity', 'Sensitivity',
               'Precision', 'AUROC', 'AUPRC', 'MCC', 'F1_score'])

    for row in data_save:
        ws.append(row.tolist())

    ws.append([''])
    ws.append(['', 'Accuracy', 'Specificity', 'Sensitivity',
               'Precision', 'AUROC', 'AUPRC', 'F1_score', 'MCC'])

    ws.append([
        'Overall Mean',
        sum(data_save[:, 1]) / loop,   
        sum(data_save[:, 2]) / loop,   
        sum(data_save[:, 3]) / loop,   
        sum(data_save[:, 4]) / loop,   
        sum(data_save[:, 5]) / loop,   
        sum(data_save[:, 6]) / loop,   
        sum(data_save[:, 8]) / loop,   
        sum(data_save[:, 7]) / loop,   
    ])

    ws.append([
        'Std Dev',
        np.std(data_save[:, 1]),
        np.std(data_save[:, 2]),
        np.std(data_save[:, 3]),
        np.std(data_save[:, 4]),
        np.std(data_save[:, 5]),
        np.std(data_save[:, 6]),
        np.std(data_save[:, 8]),
        np.std(data_save[:, 7]),
    ])

    ws.append([
        'Formatted Data',
        f"{round(sum(data_save[:, 1]) / loop * 100, 2)} ± {round(np.std(data_save[:, 1]) * 100, 2)}",
        f"{round(sum(data_save[:, 2]) / loop * 100, 2)} ± {round(np.std(data_save[:, 2]) * 100, 2)}",
        f"{round(sum(data_save[:, 3]) / loop * 100, 2)} ± {round(np.std(data_save[:, 3]) * 100, 2)}",
        f"{round(sum(data_save[:, 4]) / loop * 100, 2)} ± {round(np.std(data_save[:, 4]) * 100, 2)}",
        f"{round(sum(data_save[:, 5]) / loop, 4)} ± {round(np.std(data_save[:, 5]), 4)}",
        f"{round(sum(data_save[:, 6]) / loop, 4)} ± {round(np.std(data_save[:, 6]), 4)}",
        f"{round(sum(data_save[:, 8]) / loop, 4)} ± {round(np.std(data_save[:, 8]), 4)}",
        f"{round(sum(data_save[:, 7]) / loop * 100, 2)} ± {round(np.std(data_save[:, 7]) * 100, 2)}",
    ])

    wb.save(output_path)
    print(f'[INFO] Results saved to: {output_path}')


def save_external_results(data_save, loop, output_path, sheet_name):
    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    try:
        wb = openpyxl.load_workbook(output_path)
    except FileNotFoundError:
        wb = openpyxl.Workbook()

    wb.create_sheet(sheet_name, 0)
    ws = wb.active

    ws.append(['', 'Accuracy', 'Sensitivity', 'Precision',
               'F1_score', 'AUROC', 'AUPRC', 'MCC', 'Specificity'])

    ws.append([
        'Overall Mean',
        sum(data_save[:, 1]) / loop,
        sum(data_save[:, 2]) / loop,
        sum(data_save[:, 3]) / loop,
        sum(data_save[:, 4]) / loop,
        sum(data_save[:, 5]) / loop,
        sum(data_save[:, 6]) / loop,
        sum(data_save[:, 7]) / loop,
        sum(data_save[:, 8]) / loop,
    ])

    ws.append([
        'Std Dev',
        np.std(data_save[:, 1]),
        np.std(data_save[:, 2]),
        np.std(data_save[:, 3]),
        np.std(data_save[:, 4]),
        np.std(data_save[:, 5]),
        np.std(data_save[:, 6]),
        np.std(data_save[:, 7]),
        np.std(data_save[:, 8]),
    ])

    ws.append([
        'Formatted Data',
        f"{round(sum(data_save[:, 1]) / loop * 100, 2)} ± {round(np.std(data_save[:, 1]) * 100, 2)}",
        f"{round(sum(data_save[:, 2]) / loop, 4)} ± {round(np.std(data_save[:, 2]), 4)}",
        f"{round(sum(data_save[:, 3]) / loop, 4)} ± {round(np.std(data_save[:, 3]), 4)}",
        f"{round(sum(data_save[:, 4]) / loop, 4)} ± {round(np.std(data_save[:, 4]), 4)}",
        f"{round(sum(data_save[:, 5]) / loop, 4)} ± {round(np.std(data_save[:, 5]), 4)}",
        f"{round(sum(data_save[:, 6]) / loop, 4)} ± {round(np.std(data_save[:, 6]), 4)}",
        f"{round(sum(data_save[:, 7]) / loop, 4)} ± {round(np.std(data_save[:, 7]), 4)}",
        f"{round(sum(data_save[:, 8]) / loop, 4)} ± {round(np.std(data_save[:, 8]), 4)}",
    ])

    ws.append(['']); ws.append(['']); ws.append([''])
    ws.append(['Results of each prediction iteration:'])
    ws.append([''])
    ws.append([' ', 'Accuracy', 'Sensitivity', 'Precision',
               'F1_score', 'AUROC', 'AUPRC', 'MCC', 'Specificity'])

    for row in data_save:
        ws.append(row.tolist())

    wb.save(output_path)
    print(f'[INFO] Results saved to: {output_path}')


###############################################################################
# 1. INTERNAL VALIDATION (LOOCV)
###############################################################################

def internal_validation(args):
    print("\n" + "=" * 70)
    print("INTERNAL VALIDATION MODE (LOOCV)")
    print("=" * 70)

    print(f'[INFO] Loading data from: {args.input}')
    df = pd.read_csv(args.input)
    sample_id, protein_id, features, label, _ = data_processing(df)

    print(f'[INFO] Dataset shape   : {features.shape}')
    print(f'[INFO] Class distribution: {np.bincount(label.astype(int))}')

    protein_indices = args.protein_indices if args.protein_indices else [3, 50, 40, 36, 83]
    print(f'[INFO] Selected proteins ({len(protein_indices)}): {[protein_id[i] for i in protein_indices]}')

    features_sel = features[:, protein_indices]
    loop         = args.n_iterations
    loo          = LeaveOneOut()
    data_save    = np.zeros((loop, 9), dtype=float)

    print(f'[INFO] Running cross-validation...')
    total_start = time.time()

    for i in range(loop):
        if not args.verbose:
            print_progress_bar(i + 1, loop, prefix='Training Progress')

        ans, pred, probs = [], [], []
        for train_idx, test_idx in loo.split(features_sel):
            X_train = features_sel[train_idx]
            X_test  = features_sel[test_idx]
            y_train = label[train_idx]
            y_test  = label[test_idx]

            if args.use_smote:
                min_count = min(int(sum(y_train == 0)), int(sum(y_train == 1)))
                if min_count >= 2:
                    k = min(5, min_count - 1)
                    try:
                        sm = SMOTE(k_neighbors=k, random_state=42)
                        X_train, y_train = sm.fit_resample(X_train, y_train)
                    except:
                        pass

            model = build_knn_model(
                n_neighbors=args.n_neighbors, leaf_size=args.leaf_size,
                algorithm=args.algorithm, weights=args.weights, p=args.p,
            )
            model.fit(X_train, y_train)
            y_pred  = model.predict(X_test)
            y_proba = model.predict_proba(X_test)[:, 1]

            ans.append(y_test[0])
            pred.append(y_pred[0])
            probs.append(y_proba[0])

        ans   = np.array(ans,   dtype=int)
        pred  = np.array(pred,  dtype=int)
        probs = np.array(probs, dtype=float)

        cm = confusion_matrix(ans, pred)
        TP = cm[1, 1]; FN = cm[1, 0]; FP = cm[0, 1]; TN = cm[0, 0]

        accuracy    = (TP + TN) / (TP + FP + FN + TN)
        specificity = TN / (TN + FP)         if (TN + FP) > 0 else 0
        sensitivity = TP / (TP + FN)         if (TP + FN) > 0 else 0
        precision   = TP / (TP + FP)         if (TP + FP) > 0 else 0
        f1_score    = (2 / ((1 / precision) + (1 / sensitivity)) if precision > 0 and sensitivity > 0 else 0)
        auroc       = roc_auc_score(ans, probs)
        auprc       = average_precision_score(ans, probs)
        mcc_denom   = sqrt((TP+FP)*(TP+FN)*(TN+FP)*(TN+FN))
        mcc         = (TP*TN - FP*FN) / mcc_denom if mcc_denom > 0 else 0

        data_save[i, 0] = i + 1
        data_save[i, 1] = accuracy
        data_save[i, 2] = specificity
        data_save[i, 3] = sensitivity
        data_save[i, 4] = precision
        data_save[i, 5] = auroc
        data_save[i, 6] = auprc
        data_save[i, 7] = mcc
        data_save[i, 8] = f1_score

        if args.verbose:
            print(f'\nIteration {i+1} | Acc: {accuracy:.4f} | AUROC: {auroc:.4f}')

        if args.plot_curves and i == loop - 1:
            plot_roc_pr_curves(ans, probs, save_path=os.path.join(args.output_dir, 'internal_roc_pr.png'))

    elapsed = time.time() - total_start
    print(f'[INFO] Total time: {elapsed:.2f} s')
    print('=' * 79)
    print(f'Averages after {loop} cross-validation iterations:')
    print(f'Accuracy    = {round(sum(data_save[:,1])/loop*100,2)} ± {round(np.std(data_save[:,1])*100,2)}')
    print(f'Specificity = {round(sum(data_save[:,2])/loop*100,2)} ± {round(np.std(data_save[:,2])*100,2)}')
    print(f'Recall      = {round(sum(data_save[:,3])/loop*100,2)} ± {round(np.std(data_save[:,3])*100,2)}')
    print(f'Precision   = {round(sum(data_save[:,4])/loop*100,2)} ± {round(np.std(data_save[:,4])*100,2)}')
    print(f'AUROC       = {round(sum(data_save[:,5])/loop,4)} ± {round(np.std(data_save[:,5]),4)}')
    print(f'AUPRC       = {round(sum(data_save[:,6])/loop,4)} ± {round(np.std(data_save[:,6]),4)}')
    print(f'MCC         = {round(sum(data_save[:,7])/loop,4)} ± {round(np.std(data_save[:,7]),4)}')
    print(f'F1_score    = {round(sum(data_save[:,8])/loop*100,2)} ± {round(np.std(data_save[:,8])*100,2)}')
    print('=' * 79)

    local_time = time.strftime('%Y%m%d_%H%M%S', time.localtime())
    save_internal_results(data_save, loop, os.path.join(args.output_dir, f'Internal_Validation_{local_time}.xlsx'), 'Sheet1')


###############################################################################
# 2. EXTERNAL VALIDATION
###############################################################################

def external_validation(args):
    print("\n" + "=" * 70)
    print("EXTERNAL VALIDATION MODE")
    print("=" * 70)

    print(f'[INFO] Loading training data from: {args.train_input}')
    train_df = pd.read_csv(args.train_input)
    _, protein_id, train_features, train_label, scaler = data_processing(train_df)

    print(f'[INFO] Loading test data from: {args.test_input}')
    test_df = pd.read_csv(args.test_input)
    _, _, test_features, test_label, _ = data_processing(test_df, scaler=scaler)

    print(f'[INFO] Training set shape        : {train_features.shape}')
    print(f'[INFO] Test set shape            : {test_features.shape}')

    protein_indices = args.protein_indices if args.protein_indices else [3, 50, 40, 36, 83]
    train_sel = train_features[:, protein_indices]
    test_sel  = test_features[:,  protein_indices]

    goodbad   = np.zeros(len(test_sel), dtype=float)
    loop      = args.n_iterations
    data_save = np.zeros((loop, 9), dtype=float)

    knn = build_knn_model(
        n_neighbors=args.n_neighbors, leaf_size=args.leaf_size,
        algorithm=args.algorithm, weights=args.weights, p=args.p,
    )

    print(f'[INFO] Running external validation...')
    total_start = time.time()

    for loop_count in range(loop):
        if not args.verbose:
            print_progress_bar(loop_count + 1, loop, prefix='Prediction Progress')

        if args.use_smote:
            sm = SMOTE(k_neighbors=5, random_state=loop_count)
            X_train, y_train = sm.fit_resample(train_sel, train_label)
        else:
            X_train, y_train = train_sel, train_label

        knn.fit(X_train, y_train)
        prediction = knn.predict(test_sel)
        probs      = knn.predict_proba(test_sel)[:, 1]

        if args.plot_curves and loop_count == loop - 1:
            plot_roc_pr_curves(test_label, probs, save_path=os.path.join(args.output_dir, 'external_roc_pr.png'))

        cm = confusion_matrix(test_label, prediction)
        TP = cm[1, 1]; FP = cm[0, 1]; FN = cm[1, 0]; TN = cm[0, 0]

        accuracy    = (TP + TN) / (TP + FP + FN + TN)
        sensitivity = TP / (TP + FN) if (TP + FN) > 0 else 0
        precision   = TP / (TP + FP) if (TP + FP) > 0 else 0
        f1_score    = (2 / ((1 / precision) + (1 / sensitivity)) if precision > 0 and sensitivity > 0 else 0)
        auroc       = roc_auc_score(test_label, probs)
        auprc       = average_precision_score(test_label, probs)
        specificity = TN / (TN + FP) if (TN + FP) > 0 else 0
        mcc_denom   = sqrt((TP+FP)*(TP+FN)*(TN+FP)*(TN+FN))
        mcc         = (TP*TN - FP*FN) / mcc_denom if mcc_denom > 0 else 0

        data_save[loop_count, 0] = loop_count + 1
        data_save[loop_count, 1] = accuracy
        data_save[loop_count, 2] = sensitivity
        data_save[loop_count, 3] = precision
        data_save[loop_count, 4] = f1_score
        data_save[loop_count, 5] = auroc
        data_save[loop_count, 6] = auprc
        data_save[loop_count, 7] = mcc
        data_save[loop_count, 8] = specificity

        if args.verbose:
            print(f'\nIteration {loop_count+1} | Acc: {accuracy:.4f} | AUROC: {auroc:.4f}')

        for idx in range(len(goodbad)):
            if test_label[idx] == prediction[idx]:
                goodbad[idx] += 1

    elapsed = time.time() - total_start
    print(f'[INFO] Total time: {elapsed:.2f} s')
    print('=' * 79)
    print(f'Averages after {loop} prediction iterations:')
    print(f'Accuracy    = {round(sum(data_save[:,1])/loop*100,2)} ± {round(np.std(data_save[:,1])*100,2)}')
    print(f'Sensitivity = {round(sum(data_save[:,2])/loop,4)} ± {round(np.std(data_save[:,2]),4)}')
    print(f'Precision   = {round(sum(data_save[:,3])/loop,4)} ± {round(np.std(data_save[:,3]),4)}')
    print(f'F1_score    = {round(sum(data_save[:,4])/loop,4)} ± {round(np.std(data_save[:,4]),4)}')
    print(f'AUROC       = {round(sum(data_save[:,5])/loop,4)} ± {round(np.std(data_save[:,5]),4)}')
    print(f'AUPRC       = {round(sum(data_save[:,6])/loop,4)} ± {round(np.std(data_save[:,6]),4)}')
    print(f'MCC         = {round(sum(data_save[:,7])/loop,4)} ± {round(np.std(data_save[:,7]),4)}')
    print(f'Specificity = {round(sum(data_save[:,8])/loop,4)} ± {round(np.std(data_save[:,8]),4)}')
    print('=' * 79)

    local_time = time.strftime('%Y%m%d_%H%M%S', time.localtime())
    save_external_results(data_save, loop, os.path.join(args.output_dir, f'External_Validation_Results_{local_time}.xlsx'), 'Sheet1')


###############################################################################
# 3. FHE PREPARE (Server Side: Build Encrypted Database)
###############################################################################

def mode_prepare(args):
    print("\n" + "=" * 70)
    print("FHE PREPARE MODE [Server Side: Build Encrypted Database]")
    print("=" * 70)
    
    df = pd.read_csv(args.train_input)
    sample_id, _, features, label, _ = data_processing(df)
    protein_indices = args.protein_indices if args.protein_indices else [3, 50, 40, 36, 83]
    features_sel = features[:, protein_indices]

    if args.use_smote:
        sm = SMOTE(k_neighbors=5, random_state=42)
        features_sel, label = sm.fit_resample(features_sel, label)

    ctx = load_context(args.private_context, require_private=True)
    enc_list = []
    total = len(features_sel)
    
    for i, row in enumerate(features_sel):
        enc_list.append(ts.ckks_vector(ctx, row.tolist()).serialize())
        print_progress_bar(i + 1, total, prefix='Encrypting DB')

    payload = {
        'n_samples': len(enc_list),
        'sample_ids': sample_id.tolist(),
        'protein_indices': protein_indices,
        'n_features': len(protein_indices),
        'encrypted_samples': enc_list,
        'labels': label.tolist(),
        'has_labels': True
    }
    
    enc_path = args.output or os.path.join(args.output_dir, DEFAULT_ENC_TRAIN)
    save_bin(enc_path, payload)

    lbl_path = os.path.join(args.output_dir, DEFAULT_TRAIN_LABELS)
    with open(lbl_path, 'wb') as f:
        pickle.dump({'labels': label.tolist(), 'n_samples': len(label)}, f)


###############################################################################
# 4. FHE RUN (Client Side: Compute FHE Distances from 2 .bin files)
###############################################################################

def mode_run(args):
    print("\n" + "=" * 70)
    print("FHE RUN MODE [Client Side: Compute Homomorphic Encrypted Distances]")
    print("=" * 70)

    ctx = load_context(args.public_context, require_private=False)
    
    val_payload = load_bin(args.input)
    tr_payload = load_bin(args.encrypted_train)

    enc_queries_bytes = val_payload['encrypted_samples']
    enc_trains_bytes = tr_payload['encrypted_samples']

    enc_queries = []
    for i, b in enumerate(enc_queries_bytes):
        enc_queries.append(ts.ckks_vector_from(ctx, b))
        print_progress_bar(i + 1, len(enc_queries_bytes), prefix='Loading Samples  ')

    enc_trains = []
    for i, b in enumerate(enc_trains_bytes):
        enc_trains.append(ts.ckks_vector_from(ctx, b))
        print_progress_bar(i + 1, len(enc_trains_bytes), prefix='Loading DB       ')

    total_pairs = len(enc_queries) * len(enc_trains)
    enc_distances = []
    done = 0
    
    for q in enc_queries:
        row_dists = []
        for t in enc_trains:
            dist_enc = (q - t).square().sum()
            row_dists.append(dist_enc.serialize())
            done += 1
            print_progress_bar(done, total_pairs, prefix='FHE Distance Calc')
        enc_distances.append(row_dists)

    payload = {
        'n_query': len(enc_queries),
        'n_train': len(enc_trains),
        'query_ids': val_payload.get('sample_ids', []),
        'enc_distances': enc_distances,
        'query_labels': val_payload.get('labels'),
        'has_query_labels': val_payload.get('has_labels'),
        'train_labels': tr_payload.get('labels')
    }
    
    out_path = args.output or os.path.join(args.output_dir, 'encrypted_result.bin')
    save_bin(out_path, payload)


###############################################################################
# 5. FHE DECRYPT (Server Side: Decrypt and Predict for Client)
###############################################################################

def mode_decrypt(args):
    print("\n" + "=" * 70)
    print("FHE DECRYPT MODE [Server Side: Decrypt and Generate Report]")
    print("=" * 70)
    
    ctx = load_context(args.private_context, require_private=True)
    result = load_bin(args.encrypted_result)

    n_query = result['n_query']
    n_train = result['n_train']
    
    query_ids = result.get('query_ids', [])
    if not query_ids or len(query_ids) != n_query:
        query_ids = [f"Sample_{i+1}" for i in range(n_query)]
        
    enc_dists = result['enc_distances']
    
    if result.get('train_labels') is not None:
        train_labels = np.array(result['train_labels'])
    else:
        lbl_data = load_bin(args.train_labels)
        train_labels = np.array(lbl_data['labels'], dtype=int)

    dist_matrix = np.zeros((n_query, n_train))
    total = n_query * n_train
    done = 0

    for i in range(n_query):
        for j in range(n_train):
            enc_v = ts.ckks_vector_from(ctx, enc_dists[i][j])
            val = float(enc_v.decrypt()[0])
            dist_matrix[i, j] = max(0.0, val) 
            done += 1
            print_progress_bar(done, total, prefix='Decrypting (Private Key)')

    print("\n[INFO] Running KNN risk assessment...")
    k = args.n_neighbors
    preds, probs = [], []
    
    for i in range(n_query):
        sorted_idx = np.argsort(dist_matrix[i])[:k]
        nearest_labels = train_labels[sorted_idx]
        nearest_dists = dist_matrix[i][sorted_idx]

        if args.weights == 'distance':
            inv_w = 1.0 / np.clip(nearest_dists, 1e-10, None)
            prob = float((inv_w * nearest_labels).sum() / inv_w.sum())
        else:
            prob = float(nearest_labels.sum()) / k

        preds.append(1 if prob >= 0.5 else 0)
        probs.append(prob)

    os.makedirs(args.output_dir, exist_ok=True)
    rows = {
        'Sample_ID': list(query_ids),
        'Prediction': ['High Infection Risk' if p == 1 else 'Low Infection Risk' for p in preds],
        'Risk_Score': np.round(probs, 4),
    }
    
    csv_path = os.path.join(args.output_dir, 'Client_Report_Decrypted.csv')
    pd.DataFrame(rows).to_csv(csv_path, index=False, encoding='utf-8-sig')
    print(f"\n[SUCCESS] 🎉 Customized report generated: {csv_path}")


###############################################################################
# MAIN ENTRY
###############################################################################

def main():
    parser = argparse.ArgumentParser(
        description='PreRisk-CoV2: SARS-CoV-2 Pre-exposure Risk Assessment Framework',
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )

    parser.add_argument('--mode', type=str, required=True,
                        choices=['internal', 'external', 'prepare', 'run', 'decrypt'],
                        help='Choose mode: internal / external / prepare / run / decrypt')

    # Data
    parser.add_argument('--input', type=str, help='Input CSV or BIN')
    parser.add_argument('--train-input', type=str, help='Training CSV')
    parser.add_argument('--test-input',  type=str, help='Test CSV')

    # FHE Options
    parser.add_argument('--public-context', default=DEFAULT_PUBLIC_CTX)
    parser.add_argument('--private-context', default=DEFAULT_PRIVATE_CTX)
    parser.add_argument('--encrypted-train', default=DEFAULT_ENC_TRAIN)
    parser.add_argument('--encrypted-result', default=None)
    parser.add_argument('--train-labels', default=DEFAULT_TRAIN_LABELS)

    # Feature selection
    parser.add_argument('--protein-indices', type=int, nargs='+', default=None,
                        help='Protein indices. Default: [3,50,40,36,83]')

    # KNN hyperparameters
    parser.add_argument('--n-neighbors', type=int,  default=5)
    parser.add_argument('--leaf-size',   type=int,  default=30)
    parser.add_argument('--algorithm',   type=str,  default='auto')
    parser.add_argument('--weights',     type=str,  default='distance')
    parser.add_argument('--p',           type=int,  default=2)

    # Training options
    parser.add_argument('--use-smote',    action='store_true', default=False)
    parser.add_argument('--n-iterations', type=int, default=100)

    # Output
    parser.add_argument('--output',       type=str, default=None)
    parser.add_argument('--output-dir',   type=str, default='./results')
    parser.add_argument('--plot-curves',  action='store_true', default=False)
    parser.add_argument('--verbose',      action='store_true', default=False)

    args = parser.parse_args()
    print_banner()
    os.makedirs(args.output_dir, exist_ok=True)
    start_time = time.time()

    if args.mode == 'internal':
        if not args.input: parser.error('--input is required for internal validation')
        internal_validation(args)
    elif args.mode == 'external':
        if not args.train_input or not args.test_input:
            parser.error('--train-input and --test-input required for external')
        external_validation(args)
    elif args.mode == 'prepare':
        if not args.train_input: parser.error('--train-input is required for prepare')
        mode_prepare(args)
    elif args.mode == 'run':
        if not args.input: parser.error('--input is required for run')
        mode_run(args)
    elif args.mode == 'decrypt':
        if not args.encrypted_result: parser.error('--encrypted-result is required for decrypt')
        mode_decrypt(args)

    print("\n" + "=" * 70)
    print("PreRisk-CoV2 completed successfully!")
    print(f"Total execution time : {time.time() - start_time:.2f} s")
    print("=" * 70 + "\n")

if __name__ == '__main__':
    main()
